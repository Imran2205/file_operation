from openpyxl import Workbook


class MergeFile:
    def __init__(self, file_list=None, output_file=None):
        self.file_list = file_list
        self.output_file = output_file
        self.data_dictionary = {}

    def read_files(self):
        for i, file_name in enumerate(self.file_list):
            with open(file_name, 'r') as f:
                lines = f.readlines()
            line_no = 0
            while line_no < len(lines):
                line_mod = lines[line_no].replace('\n', '').strip().split(',')
                if len(line_mod) <= 3:
                    line_no += 1
                    continue
                if line_mod[0]:
                    if line_mod[0].lower() == 'results':
                        line_no += 1
                        continue
                    if line_mod[0].lower() == 'corners':
                        temp_line_no = line_no - 1
                        temp_line_mod = lines[temp_line_no].replace('\n', '').strip().split(',')
                    else:
                        temp_line_no = line_no
                        temp_line_mod = line_mod
                    if not line_mod[0] in self.data_dictionary.keys():
                        self.data_dictionary[line_mod[0]] = {}
                    while temp_line_mod[1]:
                        if not line_mod[0].lower() == 'corners':
                            self.data_dictionary[line_mod[0]][f'{temp_line_mod[1]}@{i}'] = \
                                temp_line_mod[2:]
                        else:
                            self.data_dictionary[line_mod[0]][f'{temp_line_mod[1]}'] = \
                                temp_line_mod[2:]
                        temp_line_no += 1
                        if temp_line_no >= len(lines):
                            break
                        temp_line_mod = lines[temp_line_no].replace('\n', '').strip().split(',')
                        if len(temp_line_mod) <= 1:
                            break
                        if temp_line_no >= len(lines):
                            break
                        temp_line_mod = lines[temp_line_no].replace('\n', '').strip().split(',')

                    line_no = temp_line_no
                line_no += 1
        # print(self.data_dictionary)
        self.gen_excel()

    def column_letter(self, num):
        col_let = ''
        div = num
        while div > 0:
            rem = div % 26
            div = div // 26
            if rem != 0:
                col_let = col_let + chr(rem + 64)
            else:
                col_let = col_let + 'Z'
        return col_let

    def gen_excel(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = 'main'
        row_count = 1
        col_count = 1
        max_width = 0
        max_width_2 = 0
        for key in self.data_dictionary.keys():
            ws1.cell(column=col_count, row=row_count, value=key)
            if len(key) + 2 > max_width_2:
                max_width_2 = len(key) + 2
                ws1.column_dimensions[self.column_letter(col_count)].width = max_width_2
            col_count += 1
            level_2_keys = list(self.data_dictionary[key].keys())
            if key.lower() != 'corners':
                level_2_keys.sort()
            for key2 in level_2_keys:
                ws1.cell(column=col_count, row=row_count, value=key2)
                if len(key2) + 2 > max_width:
                    max_width = len(key2) + 2
                    ws1.column_dimensions[self.column_letter(col_count)].width = max_width
                col_count += 1
                for val in self.data_dictionary[key][key2]:
                    try:
                        if 'inf' not in val:
                            ws1.cell(column=col_count, row=row_count, value=float(val))
                        else:
                            ws1.cell(column=col_count, row=row_count, value=val)
                    except:
                        ws1.cell(column=col_count, row=row_count, value=val)
                    col_count += 1
                row_count += 1
                col_count = 2
            row_count += 1
            col_count = 1
        wb.save(self.output_file)


if __name__ == '__main__':
    files = [
        './script_related_files/bench_lpsn.csv',
        './script_related_files/bench_lvt.csv',
        './script_related_files/bench_spln.csv'
    ]
    out_file_name = 'output.xlsx'
    m_f_obj = MergeFile(file_list=files, output_file=out_file_name)
    m_f_obj.read_files()
